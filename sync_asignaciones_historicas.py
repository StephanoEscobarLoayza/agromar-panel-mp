"""
Rellena asignaciones historicas (lote -> corrida) leyendo el detalle de
lotes de cada hoja de corrida ya cerrada en Trazabilidad. `sync_corridas.py`
solo trae la CABECERA (MP kg, fechas, rendimiento); sin este script el
kg_saldo de los lotes viejos no reflejaba lo realmente consumido, y un lote
marcado PROCESADO en el Sheet de recepcion aparecia con saldo completo en
la app (bug detectado por Stephano).

Regla de seguridad: por cada par (lote, corrida) SOLO se inserta si no
existe TODAVIA ninguna asignacion para ese par (sin importar el origen).
Nunca se toca ni se ajusta una asignacion ya registrada - lo que el
personal ya cargo por la app queda como fuente de verdad. Esto tambien
hace el script idempotente: correrlo de nuevo no duplica nada.

La columna autoritativa de kg consumidos es "Peso con descuento" (no
"Peso Neto"): se verifico que su suma por hoja coincide exactamente con
el "MP kg" que ya trae cada corrida (ej. 29-agosto JCC: suma de la
columna = 63137 = MP kg de esa hoja).

Uso:
    python sync_asignaciones_historicas.py "C:\\ruta\\a\\Trazabilidad Nar 2026.xlsx"
"""
import os
import sys
import unicodedata
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env")

DEFAULT_PATH = r"C:\Users\Stephano\OneDrive\Desktop\Trazabilidad Nar 2026 (1).xlsx"

DATABASE_URL = os.environ.get("DATABASE_URL_POOLED") or os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit("Falta DATABASE_URL en .env")

engine = create_engine(DATABASE_URL)

MARCADOR = "Importado de Trazabilidad (detalle historico)"


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def encontrar_header_row(ws, max_row=12, max_col=8):
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            if norm(ws.cell(row=r, column=c).value) == "lote":
                return r
    return None


def encontrar_columna(ws, header_row, variantes, max_col=20):
    for c in range(1, max_col + 1):
        if norm(ws.cell(row=header_row, column=c).value) in variantes:
            return c
    return None


def tipo_almacen_origen(valor):
    v = norm(valor)
    if not v:
        return None
    if v[0] == "s":
        return "SILO"
    if v[0] == "b":
        return "BINES"
    return None


def leer_detalle_lotes(ws):
    """Devuelve {lote_numero: {kg, origen}}, sumando si un lote aparece mas
    de una vez en la misma hoja."""
    header_row = encontrar_header_row(ws)
    if header_row is None:
        return {}

    col_lote = encontrar_columna(ws, header_row, {"lote"})
    col_descuento = encontrar_columna(ws, header_row, {"peso con descuento"})
    col_neto = encontrar_columna(ws, header_row, {"peso neto"})
    col_origen = encontrar_columna(ws, header_row, {"silo / bines", "silo/bines", "silo - bines"})

    if col_lote is None or (col_descuento is None and col_neto is None):
        return {}

    detalle = {}
    r = header_row + 1
    while True:
        lote_val = ws.cell(row=r, column=col_lote).value
        if not isinstance(lote_val, (int, float)) or lote_val <= 0:
            break
        # "Peso con descuento" es la cifra autoritativa (coincide con el "MP
        # kg" de la hoja) cuando esta presente; si esa hoja no aplico
        # descuento por brix la deja en blanco, y ahi se usa "Peso Neto".
        peso_val = ws.cell(row=r, column=col_descuento).value if col_descuento else None
        if not isinstance(peso_val, (int, float)) or peso_val <= 0:
            peso_val = ws.cell(row=r, column=col_neto).value if col_neto else None
        if isinstance(peso_val, (int, float)) and peso_val > 0:
            lote_numero = int(lote_val)
            origen = tipo_almacen_origen(ws.cell(row=r, column=col_origen).value) if col_origen else None
            if lote_numero in detalle:
                detalle[lote_numero]["kg"] += float(peso_val)
            else:
                detalle[lote_numero] = {"kg": float(peso_val), "origen": origen}
        r += 1
    return detalle


def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    if not os.path.exists(ruta):
        raise SystemExit(f"No se encontro el archivo: {ruta}")

    wb = openpyxl.load_workbook(ruta, data_only=True)

    insertados = 0
    ya_registrado = 0
    lote_desconocido = 0
    rechazados = 0

    with engine.connect() as conn:
        corridas = {
            row.nombre: (row.id, row.fecha_inicio)
            for row in conn.execute(text("SELECT id, nombre, fecha_inicio FROM corridas"))
        }
        lotes_conocidos = {row.numero for row in conn.execute(text("SELECT numero FROM lotes"))}

    for nombre_hoja in wb.sheetnames:
        if "resumen" in nombre_hoja.lower():
            continue
        nombre = nombre_hoja.strip()
        if nombre not in corridas:
            continue
        corrida_id, fecha_inicio = corridas[nombre]

        ws = wb[nombre_hoja]
        detalle = leer_detalle_lotes(ws)

        for lote_numero, info in detalle.items():
            if lote_numero not in lotes_conocidos:
                lote_desconocido += 1
                continue

            with engine.connect() as conn:
                existe = conn.execute(
                    text("SELECT 1 FROM asignaciones WHERE lote_numero = :l AND corrida_id = :c"),
                    {"l": lote_numero, "c": corrida_id},
                ).fetchone()
            if existe:
                ya_registrado += 1
                continue

            try:
                with engine.begin() as conn:
                    conn.execute(
                        text(
                            """
                            INSERT INTO asignaciones
                                (lote_numero, corrida_id, fecha_proceso, tipo_almacen_origen, kg_asignados, observaciones)
                            VALUES
                                (:lote_numero, :corrida_id, :fecha_proceso, :origen, :kg, :obs)
                            """
                        ),
                        {
                            "lote_numero": lote_numero,
                            "corrida_id": corrida_id,
                            "fecha_proceso": fecha_inicio,
                            "origen": info["origen"],
                            "kg": info["kg"],
                            "obs": MARCADOR,
                        },
                    )
                insertados += 1
            except Exception as e:
                rechazados += 1
                print(f"  [!] {nombre} / lote {lote_numero}: {str(e).splitlines()[0]}")

    print(
        f"\nListo: {insertados} asignaciones creadas, "
        f"{ya_registrado} ya tenian registro (no se tocaron), "
        f"{lote_desconocido} con lote desconocido, "
        f"{rechazados} rechazadas por saldo insuficiente."
    )


if __name__ == "__main__":
    main()
