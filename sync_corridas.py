"""
Trae las corridas desde "Trazabilidad Nar 2026.xlsx" (una hoja por corrida)
hacia la tabla `corridas` de la base de datos.

Solo importa la CABECERA de cada corrida (nombre, tipo de proceso, fechas,
MP kg / rendimiento / brix ya documentados) - no los lotes individuales que
alimentaron cada corrida (eso ya se maneja con las asignaciones registradas
en la app, o con la reconciliacion manual que ya se hizo para las corridas
mas viejas).

Todas las corridas que trae este script quedan como estado='cerrada' -
Trazabilidad solo documenta corridas que YA PASARON (el flujo real es al
reves de lo que uno pensaria: primero se registra el consumo en la app,
DESPUES se arma la hoja de Trazabilidad con esos numeros - ver README).
Las corridas nuevas/en curso se crean directo desde la app, no por aqui.

Se puede correr las veces que haga falta: hace upsert por `nombre` (el
nombre de la hoja, tal cual) - las corridas que ya existen se actualizan
(por si se corrigio algun numero en Trazabilidad), las nuevas se agregan.

Uso:
    python sync_corridas.py "C:\\ruta\\a\\Trazabilidad Nar 2026.xlsx"

Si no se pasa la ruta, usa por defecto la del Desktop del usuario.
"""
import os
import sys
import unicodedata
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from sqlalchemy import create_engine, text

BASE_DIR = Path(__file__).parent
load_dotenv(BASE_DIR / ".env")

DEFAULT_PATH = r"C:\Users\Stephano\OneDrive\Desktop\Trazabilidad Nar 2026 (1).xlsx"

DATABASE_URL = os.environ.get("DATABASE_URL_POOLED") or os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    raise SystemExit("Falta DATABASE_URL en .env")

engine = create_engine(DATABASE_URL)

TIPOS_PROCESO = ["ICEGEN", "JSA", "JCC", "JSC"]  # orden: los mas especificos primero

ETIQUETAS = {
    "mp_kg_objetivo": ["mp kg"],
    "rendimiento": ["rendimiento"],
    "brix_promedio_tk": ["brix promedio tk"],
}


def norm(s):
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode("ascii")
    return s.strip().lower()


def detectar_tipo_proceso(nombre_hoja):
    up = nombre_hoja.upper()
    for tipo in TIPOS_PROCESO:
        if tipo in up:
            return tipo
    return None


def leer_fechas(ws):
    """Hojas nuevas: D1='INICIO', E1='FINAL', fechas en D2/E2.
    Hojas viejas: solo 'FECHA DE PROCESO:' en B2, fecha unica en D2 - se usa
    la misma fecha para inicio y final (es lo unico que se sabe)."""
    if norm(ws["D1"].value) == "inicio":
        inicio = ws["D2"].value
        final = ws["E2"].value
        return inicio, (final or inicio)
    else:
        inicio = ws["D2"].value
        return inicio, inicio


def leer_etiquetas(ws, max_row=45):
    """Busca en la columna B (filas 1..max_row) las etiquetas conocidas y
    devuelve el valor de la columna G de esa misma fila."""
    encontrados = {}
    for row in range(1, max_row + 1):
        etiqueta = norm(ws.cell(row=row, column=2).value)  # columna B
        if not etiqueta:
            continue
        for campo, variantes in ETIQUETAS.items():
            if campo in encontrados:
                continue
            if etiqueta in variantes:
                valor = ws.cell(row=row, column=7).value  # columna G
                if isinstance(valor, (int, float)):
                    encontrados[campo] = float(valor)
    return encontrados


def main():
    ruta = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    if not os.path.exists(ruta):
        raise SystemExit(f"No se encontró el archivo: {ruta}")

    wb = openpyxl.load_workbook(ruta, data_only=True)

    insertados = actualizados = omitidos = 0

    with engine.begin() as conn:
        for nombre_hoja in wb.sheetnames:
            if "resumen" in nombre_hoja.lower():
                continue

            ws = wb[nombre_hoja]
            nombre = nombre_hoja.strip()

            inicio, final = leer_fechas(ws)
            if inicio is None:
                print(f"  omitida (sin fecha): {nombre}")
                omitidos += 1
                continue

            tipo_proceso = detectar_tipo_proceso(nombre_hoja)
            etiquetas = leer_etiquetas(ws)

            data = {
                "nombre": nombre,
                "tipo_proceso": tipo_proceso,
                "fecha_inicio": inicio,
                "fecha_final": final,
                "mp_kg_objetivo": etiquetas.get("mp_kg_objetivo"),
                "rendimiento": etiquetas.get("rendimiento"),
                "brix_promedio_tk": etiquetas.get("brix_promedio_tk"),
            }

            resultado = conn.execute(
                text(
                    """
                    INSERT INTO corridas
                        (nombre, tipo_proceso, fecha_inicio, fecha_final, mp_kg_objetivo,
                         rendimiento, brix_promedio_tk, estado)
                    VALUES
                        (:nombre, :tipo_proceso, :fecha_inicio, :fecha_final, :mp_kg_objetivo,
                         :rendimiento, :brix_promedio_tk, 'cerrada')
                    ON CONFLICT (nombre) DO UPDATE SET
                        tipo_proceso = EXCLUDED.tipo_proceso,
                        fecha_inicio = EXCLUDED.fecha_inicio,
                        fecha_final = EXCLUDED.fecha_final,
                        mp_kg_objetivo = EXCLUDED.mp_kg_objetivo,
                        rendimiento = EXCLUDED.rendimiento,
                        brix_promedio_tk = EXCLUDED.brix_promedio_tk,
                        estado = 'cerrada'
                    RETURNING (xmax = 0) AS es_insercion
                    """
                ),
                data,
            )
            if resultado.scalar():
                insertados += 1
            else:
                actualizados += 1

    print(f"\nListo: {insertados} corridas nuevas, {actualizados} actualizadas, {omitidos} omitidas.")


if __name__ == "__main__":
    main()
